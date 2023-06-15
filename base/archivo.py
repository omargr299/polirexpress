
from openpyxl import Workbook 
from datetime import datetime

def excel(datos:list):
    wb = Workbook()
    sheet = wb.active
    if sheet is None: return
    sheet.cell(row=1,column=1,value='ID')
    sheet.cell(row=1,column=2,value='Destino')
    sheet.cell(row=1,column=3,value='Fecha')
    sheet.cell(row=1,column=3,value='Hora')
    for indice,registro in enumerate(datos):
        sheet.cell(row=indice+2,column=1,value=registro[0])
        sheet.cell(row=indice+2,column=2,value=registro[1])
        sheet.cell(row=indice+2,column=3,value=registro[2].strftime('%d/%m/%Y'))
        sheet.cell(row=indice+2,column=3,value=registro[2].strftime('%H:%M:%S'))
    date = datetime.now().strftime('%d-%m-%Y %H-%M')
    wb.save(f'registros {date}.xlsx')
    print('Archivo guardado')
    